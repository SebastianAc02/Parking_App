from dash import Dash 
from dash import dcc
from dash import html
from dash.dependencies import Input, Output, State
import os
import pandas as pd
import plotly.graph_objects as go
import numpy as np
from PIL import Image
from dash.dash import no_update





from datetime import datetime
import openpyxl

EXCEL_PATH = "./carparkinglot.xlsx"
 
class ExcelColumns:
    Type = None
    Plate = None
    Entrance_Date = None
    Exit_Date = None
    Cell = None
    Cell_State = None 


class Parking_Lot:
    def  __init__(self,max_opcupation):
        self.max_ocupation = max_opcupation
        self.current_ocupation = 0 
        #column = {"Type", "Plate","Entrance_Date", "Exit_Date", "Cell", "Cell_State" }
        self.parking_lot = self.createParkingLot()
        
        self.Excel = Excel(ExcelColumns)
        self.getVehicles()
        print("sucess")

    def getParkingCurrent(self):
        return self.parking_lot
    
    def addVehicle(self, type, plate , cell ):
        if(self.current_ocupation >= self.max_ocupation):
            return False

        if(self.Excel.get_car_by_plate(plate, "Occupied")):
            return False 
        
        if not(self.Excel.checkCellNUmberisFree(cell)):
            return False



        self.current_ocupation = self.current_ocupation+1
        newVehicle =  Vehicle(type, plate, cell )
       
        self.Excel.open_car(newVehicle)

        fila = (num - 1) // self.columns
        columna = (num - 1) % self.columns 
        print('Hi thereee',fila, columna)
        self.parking_lot[fila][columna] = "Occupied"
         
        return True 

    def getCols(self):
        return self.columns 

    def getRows(self):
        return self.rows

    def createParkingLot(self):
        self.rows = 10
        self.columns = 8
        parking_lot = [['Empty' for _ in range(self.columns)] for _ in range(self.rows)]
        return parking_lot
        

    def exitVehicle(self, plate):
      
        if not (self.Excel.get_car_by_plate(plate, "Occupied") ):
            return False 

        self.current_ocupation = self.current_ocupation-1
        self.Excel.close_car(plate)

    def isVehicleRegistered(self):
        print("heloo world") 
        
    def ifFull(self):
        return self.current_ocupation >= self.max_ocupation 

    def getVehicles(self):
        cars = self.Excel.getCarsByState()

        if (len(cars) <= 0):
            return []
        
        for num in cars:
            print(f"cars: for {num}")
            
            fila = (num - 1) // self.columns
            columna = (num - 1) % self.columns 
            self.parking_lot[fila][columna] = "Occupied"

        
        

    


class Vehicle:
    def __init__(self, type, plate, cell ):
        if not (self.checkPlateValid(plate)):
             raise ValueError("La placa no es válida")
        if not(cell >= 0 and cell <= 80):
            raise ValueError("La celda no es válida")

        self.type = type
        self.plate = plate
        self.entryHour = self.getHourNow()
        self.exitDate = ""
        self.cell = cell 
        self.setCellState("Occupied")

    def setCellState(self, cellState):
        self.cellState = cellState

    def checkPlateValid(self, plate):
        plate = plate.replace(" ", "").upper()

        if len(plate) != 6:
            return False

        num_letters = sum(c.isalpha() for c in plate)
        num_numbers = sum(c.isdigit() for c in plate)

        # Comprobar si la placa tiene exactamente 3 letras y 3 números
        if num_letters == 3 and num_numbers == 3:
            return True
        else:
            return False

    def getHourNow(self):
        ahora = datetime.now()
        fecha_hora = ahora.strftime("%Y-%m-%d %H:%M:%S")
        return fecha_hora






class ServerGPT:

    def __init__(self, name, app):
        self.app = app
        self.app_name = name
        self.file_to_watch = EXCEL_PATH
        self.layout = self.create_layout()
        self.Parking = Parking_Lot(80)
        self.set_callbacks()

    
    def create_graph(self, parking_lot, flag=0 ):
        # rows = 5
        # columns = 8
        # parking_lot = [['Empty' for _ in range(columns)] for _ in range(rows)]
         

        # # Example of adding a car at row 2, column 4
        # parking_lot[0][0] = "Occupied"
       #rint(f"parking lot: {parking_lot}")
        fig = go.Figure()
         
        images = []  # List to hold the images

        for i in range(5):
            for j in range(8):
                color = None 
                
                if(parking_lot[i][j] == "Empty"):
                    color="white"
                else:
                    color="lightblue"
                 
                fig.add_trace(go.Scatter(
                    x=[j+1], y=[i+1],
                    mode='markers+text',
                    marker=dict(
                        size=50, 
                        color=color,
                        line=dict(
                            color='Blue',
                            width=2
                        ),
                        symbol='square'
                    ),

                   text = str((i * 8 + j + 1) + (20*flag)) if parking_lot[i][j] == "Empty" else "🏎️",
                    textposition="middle center",
                    hovertemplate = f'Row: {i+1}<br>Column: {j+1}<br>Status: {"Empty" if parking_lot[i][j] == "Empty" else "Occupied"}<extra></extra>'
                ))
               

        fig.update_layout(showlegend=False,
                        xaxis=dict(range=[0, 10], showgrid=False, zeroline=False, showticklabels=False),
                        yaxis=dict(range=[0, 7], showgrid=False, zeroline=False, autorange='reversed', showticklabels=False),
                        plot_bgcolor='White',
                        images=images)

        return fig


    def create_layout(self):
       
        return  html.Div(children=[
        # First row with H1
        html.Div([
            html.H1(f"{self.app_name}", style={'textAlign': 'center'})
        ]),
        # Next row with a dropdown and a button
        html.Div([
            dcc.Dropdown(
                id='select-parking-space',
                options=[
                    {'label': 'A', 'value': 'A'},
                    {'label': 'B', 'value': 'B'}
                ],
                value='A',
                style={'width':'60%', 'text-align':'center'}
            ),
            html.Button('Check Statistics', id='button-check-statistics', style={'width':'40%'}),
        ], style={'display':'flex', "justify-content":"space-between"}),
        # Graph placeholder
        html.Div([
            
            dcc.Graph(
                id='graphA',
                figure= {}
            )
        ], id="DivSectionA",
         style={'display':'none'} 
         ),
        html.Div([
            dcc.Graph(
                id='graphB',
                figure= {}
            )
        ], id="DivSectionB",
        style={'display':'none'}),
        # Div with a button and two inputs
        html.Div([
          
            html.Div([
                html.Label('Plate:'),
                dcc.Input(id='plate-input', type='text'),
                html.Label('Space:'),
                dcc.Input(id='space-input', type='number'),
                html.Button('Ok', id='submitbtn', n_clicks=0),
            ], style={'display': 'flex', 'justifyContent': 'space-between'})
            ]),
            html.Div(id="DummyOut", children={})
            
           
        ])
       

    def set_callbacks(self):

        
        @self.app.callback(
            Output(component_id='DummyOut',component_property='children'),
            [Input(component_id="submitbtn", component_property="n_clicks")],
            [State(component_id="plate-input", component_property="value"),
            State(component_id="space-input", component_property="value"),
            State(component_id="select-parking-space", component_property="value")],
            prevent_initial_call= True 
        )

        def update_graph2(n, plate, place,parkingSection  ):
             
            if(n>=0 and plate  != "" and place != ""):
                if not (self.Parking.addVehicle('Car', plate, place)):
                    print("error adding a new car, identify the error to display it in ui")
                else:
                    print("working just fine ")
            else:
                pass 

            update_graph(parkingSection)

            return html.Div()

            
        
         # Callback to update graph baseed on input section A/ B 
        @self.app.callback(
        [Output(component_id='graphA',component_property='figure'),
        Output(component_id='DivSectionA',component_property='style'),
        Output(component_id='graphB',component_property='figure'),
        Output(component_id='DivSectionB',component_property='style')],
        [Input(component_id="select-parking-space", component_property="value")],
        
        prevent_initial_call= False 
       )

        def update_graph( value_space ):

            print(f"value space: {value_space}")
        
            matrix = self.Parking.getParkingCurrent()
        
            if(value_space == 'A' ):
            
                parking_lot = matrix[:5] 
                
                figure = self.create_graph(parking_lot)
                return  figure , {"display":""}, no_update, {"display":"none"}
            else:
                
                parking_lot = matrix[-5:]
                
                figure = self.create_graph(parking_lot, 2)
                return no_update ,  {"display":"none"}, figure,{"display":""}
    
        

    def run(self, debug=False):
        self.app.layout = self.layout
        self.app.run_server(debug=debug)

class Excel:


    def __init__(self, HEAD):
       
        self.excel_path = EXCEL_PATH  # Reemplaza con la ruta real de tu archivo Excel
        self.excel = openpyxl.load_workbook(self.excel_path)
        self.sheet = self.excel.active
        
        self.headers = [attr for attr in HEAD.__dict__.keys() if not attr.startswith('__')]


         
        existing_headers = self.sheet[1]
        if existing_headers[0].value is None:
            for idx, header in enumerate(self.headers, start=1):
                self.sheet.cell(row=1, column=idx, value=header)
        
            self.excel.save(self.excel_path)
            
        else:
            return
        
        
        

    def open_car(self, obj):
        row_data= list(vars(obj).values())
        print("ola")
        print(f"data: {row_data}")
        self.sheet.append(row_data)
        self.excel.save(self.excel_path)

    def close_car(self, plate):

        for i, row in enumerate(self.sheet.iter_rows(min_row=2, values_only=True), start=2):
            print(row[1])
            if row and row[1] == plate:
                # Editar la columna 5 y reemplazar el texto existente por "Empty"
                print(self.sheet.cell)
                self.sheet.cell(row=i, column=6, value="Empty")

                # Escribir la fecha actual y la hora en la columna 3
                current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                self.sheet.cell(row=i, column=4, value=current_datetime)

                self.sheet.cell(row=i, column= 5, value='none')

                self.excel.save(self.excel_path)
                break
            else:
                print("plate was not found")
                    

    def checkCellNUmberisFree(self, cellNumber, cellState="Occupied"):
         for row in self.sheet.iter_rows(min_row=2 , values_only=True):
            if row and row[4] == cellNumber and row[5] == cellState :
                return False
        
         return True 


    def get_car_by_plate(self, plate, cellState ):

        
            for row in self.sheet.iter_rows(min_row=2 , values_only=True):
                if row and row[1] == plate and row[5] == cellState :
                    car_data = {}
                    for i, header in enumerate(self.headers):
                        car_data[header] = row[i]
                    return car_data
    
    def getCarsByState(self, cellState="Occupied", ):

        cars = []
        for row in self.sheet.iter_rows(min_row=2 , values_only=True):
            if row and row[5] == "Occupied" :
                cars.append(row[4])

        return cars 


    def is_car_parked(self, plate, cellState="Occupied"):
         for row in self.sheet.iter_rows(min_row=2, values_only=True):
            if row and row[1] == plate and row[5] == cellState :
                return true 
         return false


#park = Parking_Lot(80)

#ark.addVehicle('Car', 'V4k3K9', 20)

#print(park.getParkingCurrent())


# Create a Dash application object
app = Dash(__name__)

# Create an instance of your Server class
server = ServerGPT("My Parking Lot App", app)

# Run the Dash application
server.run(debug=True)
